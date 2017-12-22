#!/usr/bin/env python
# -- coding: utf-8 -*-
'''
Created on 2016年7月4日

@author: Wan
'''
import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy.sql.expression import and_

from common import module_path
from database import (BuliangFangan, BuliangKuchun, CangkuLiushui,
                      ProductClassification, ProductLiushui, db_session,
                      reset_table)
from win32office import Excel

data_dir = os.path.join(module_path(), 'data')


def init_buliang_fangan():
    reset_table('buliang_fangan')
    wb_file = os.path.join(data_dir, '不良品处理对照表.xlsx')
    wb = load_workbook(wb_file)
    ws = wb.get_sheet_by_name('对照表')
    for row in ws.iter_rows('A2:C{}'.format(ws.max_row)):
        product_name, buliang_name, bili = row
        fangan = BuliangFangan(product_name=product_name.value,
                               buliang_name=buliang_name.value,
                               chuliliang=bili.value)
        db_session.add(fangan)
    db_session.commit()


def chuli_liushuizhang(file_path):
    excel = Excel(file_path)
    excel.select('流水账')
    for row in range(3, excel.max_row() + 1):
        leixing = excel.get_cell_value(row, 1)
        if leixing == '往来销售':
            code = excel.get_cell_value(row, 5)
            name = excel.get_cell_value(row, 6)
            if name:
                name = _strip_name(name)
                print(name)


def _strip_name(str_):
    str_ = re.sub(r'[\u4e00-\u9fa5]+', '', str_)  # 去掉中文
    str_ = re.sub(r'HD[0-9]{1,2}', "", str_)  # 去掉HD01~HD99
    str_ = re.sub(r'0[1-9]$', "", str_)  # 去掉 01~09后缀
    str_ = re.sub(r'[1-9]0CP$', "", str_)  # 去掉1~90CP
    str_ = re.sub(r'[1-9]\.[1-9]CP$', "", str_)  # 去掉1.1CP~9.9CP
    str_ = re.sub(r'[0-9]{3}PS$', "", str_)  # 去掉100PS~999PS
    str_ = re.sub(r'[0-9]{2}PS$', "", str_)  # 去掉10PS~99PS
    return str_


def _fenci(str_):
    with open('fenci.dic', 'r') as fp:
        d = set(fp.readlines()).reverse()
        for item in d:
            if str.find(item) >= 0:
                key = item


def load_product_class(file_path, unmerge=False):
    reset_table('product_classification')
    excel = Excel(file_path)
    if unmerge:
        _unmerge_product_class(excel)  # unmerge and select 2th sheet
    excel.select(2)
    for row in range(2, excel.max_row() + 1):
        (xuhao, new_costing_classification, costing_classification,
         accounting_classification, slug, model_name,
         unit, product_code, new_note,
         note, people_name, people_code) = [excel.get_cell_value(row=row, col=col) for col in range(1, 13)]

        if not slug:
            continue
        if product_code:
            product_code = product_code.strip()
            new_code = "'%s" % product_code[:5]
        slug = _strip(slug)
        model_name = _strip(model_name)

        if isinstance(new_note, str):
            new_note = new_note.replace("新增", "").replace(" ", "")
            m = re.match(r'^(\d+)[\.\u5e74]{1}(\d+)[\.\u6708]{1}(\d+)[\u65e5]?$', new_note)
            if m:
                year, month, day = m.groups()
                new_note = datetime(int(year), int(month), int(day))

        p = ProductClassification(accounting_classification=accounting_classification,
                                  costing_classification=costing_classification,
                                  new_costing_classification=new_costing_classification,
                                  model_name=model_name,
                                  slug=slug,
                                  unit=unit,
                                  product_code=product_code,
                                  note=note,
                                  create_time=new_note,
                                  people_name=people_name)
        db_session.add(p)
        excel.set_range_value(1, row, [[xuhao, new_costing_classification, costing_classification,
                                        accounting_classification, slug, model_name,
                                        unit, product_code, datetime.strftime(
                                            new_note, '%Y-%m-%d') if new_note else None,
                                        note, people_name, people_code, new_code]])
        print('add %s' % row)
    db_session.commit()
    excel.save()
    excel.quit()


def _strip(some_string):
    if not some_string:
        return ''
    some_string = str(some_string).replace("）", ")").replace("（", "(").strip()
    pattern = re.compile(r'\s+')
    some_string, number = pattern.subn(' ', some_string)  # 把连续的空格替换成单一空格
    return some_string


def _unmerge_product_class(excel):
    excel.select(1)
    data = []
    for row in range(5, excel.max_row() - 10):
        row_data = []
        for col in range(1, 13):
            cell = excel.ws.Cells(row, col)
            _range = excel.ws.Range(cell.Address).Mergearea
            if _range.Address != cell.Address:   # 合并区域
                row_data.append(_range.Value[0][0])
            else:
                row_data.append(cell.Value)
        data.append(row_data)
    excel.select(2)
    excel.set_range_value(1, 1, data)


def load_buliang_kuchun(file_path, start_row, stop_row):
    reset_table('buliang_kuchun')
    excel = Excel(file_path)
    excel.select()
    for row in range(start_row, stop_row + 1):
        product_name = excel.get_cell_value(row=row, col=1)
        amount = excel.get_cell_value(row=row, col=2)
        kuchun = BuliangKuchun(product_name=product_name, amount=amount)
        db_session.add(kuchun)
        print('%s \t %s' % (product_name, amount))
    db_session.commit()
    excel.quit()


def load_cangku_liushui(file_path, start_row, stop_row):
    # reset_table('cangku_liushui')
    excel = Excel(file_path)
    excel.select('流水账')
    for row in range(start_row, stop_row + 1):
        yewu_type, jilu_date, danhao, kehu, chanpin_bianma,\
            product_name, spec, batch, amount, product_date,\
            note, peifang_version = [
                excel.get_cell_value(row=row, col=col) for col in range(1, 13)]
        if not isinstance(jilu_date, datetime):
            jilu_date = None
        if not isinstance(product_date, datetime):
            product_date = None
        liushui = CangkuLiushui(yewu_type=yewu_type,
                                jilu_date=jilu_date,
                                danhao=danhao,
                                kehu=kehu,
                                chanpin_bianma=chanpin_bianma,
                                product_name=product_name,
                                spec=spec,
                                batch=batch.strip('.0'),
                                amount=amount,
                                product_date=product_date,
                                peifang_version=peifang_version,
                                note=note)
        db_session.add(liushui)
        print('add %s \t %s \t %s \t %s' % (yewu_type, jilu_date, product_name, batch))
    db_session.commit()
    excel.quit()


def load_product_liushui(file_path, start_row, stop_row):
    excel = Excel(file_path)
    excel.select()
    for row in range(start_row, stop_row + 1):
        (kind, product_date, product_name, batch,
         ji_hua_zhong, pei_liao_liang, he_zhong_liang,
         yan_mo_hou, yan_mo_sun_hao, jia_liao_liang,
         san_lei, fan_hui_you, jia_liao_hou,
         sheng_yu_you, guan_shu, gui_ge, ru_ku_liang, gui_ge_2, gu_hua_ji,
         bao_zhuang_sun_hao, zong_sun_hao, sun_hao_lv,
         wan_cheng_ri) = [excel.get_cell_value(row=row, col=col) for col in range(1, 24)]
        if not isinstance(product_date, datetime):
            product_date = None
        if not isinstance(wan_cheng_ri, datetime):
            wan_cheng_ri = None
        liushui = ProductLiushui(kind=kind,
                                 product_date=product_date,
                                 product_name=product_name,
                                 batch=batch.strip(".0"),
                                 ji_hua_zhong=ji_hua_zhong,
                                 pei_liao_liang=pei_liao_liang,
                                 he_zhong_liang=he_zhong_liang,
                                 yan_mo_hou=yan_mo_hou,
                                 yan_mo_sun_hao=yan_mo_sun_hao,
                                 jia_liao_liang=jia_liao_liang,
                                 san_lei=san_lei,
                                 fan_hui_you=fan_hui_you,
                                 jia_liao_hou=jia_liao_hou,
                                 sheng_yu_you=sheng_yu_you,
                                 guan_shu=guan_shu,
                                 gui_ge=gui_ge,
                                 ru_ku_liang=ru_ku_liang,
                                 gui_ge_2=gui_ge_2,
                                 gu_hua_ji=gu_hua_ji,
                                 bao_zhuang_sun_hao=bao_zhuang_sun_hao,
                                 zong_sun_hao=zong_sun_hao,
                                 sun_hao_lv=sun_hao_lv,
                                 wan_cheng_ri=wan_cheng_ri)
        db_session.add(liushui)
        print('%s \t %s \t %s' % (wan_cheng_ri, product_name, batch))
    db_session.commit()
    excel.quit()


def jisuan_ruku_liushui(out_file):
    excel = Excel(out_file)
    excel.select('成品入库流水')
    liushui = CangkuLiushui.query.filter(and_(CangkuLiushui.yewu_type == "产品进仓",
                                              CangkuLiushui.jilu_date >= datetime(2016, 5, 31),
                                              CangkuLiushui.jilu_date <= datetime(2016, 6, 30))).all()
    buliangfang = BuliangFangan.query.all()
    data = []
    for batch in liushui:
        flag_find = False
        jilu_date = datetime.strftime(batch.jilu_date, '%Y-%m-%d')
        for fangan in buliangfang:
            if batch.product_name.find(fangan.product_name) >= 0:
                flag_find = True
                data.append([jilu_date, batch.product_name, batch.batch, batch.amount,
                             fangan.buliang_name, fangan.chuliliang])
                break
        if not flag_find:
            # 兑入同型号10%
            data.append([jilu_date, batch.product_name, batch.batch, batch.amount,
                         '理论可处理同型号10%', 0.1])
    excel.set_range_value(1, 2, data)
    excel.save()


def jisuan_product_liushui(out_file):
    excel = Excel(out_file)
    excel.select('生产流水')
    liushui = ProductLiushui.query.filter(and_(ProductLiushui.wan_cheng_ri >= datetime(2016, 5, 31),
                                               ProductLiushui.wan_cheng_ri <= datetime(2016, 6, 30))).all()
    buliangfang = BuliangFangan.query.all()
    data = []
    for batch in liushui:
        flag_find = False
        wan_cheng_ri = datetime.strftime(batch.wan_cheng_ri, '%Y-%m-%d')
        for fangan in buliangfang:
            if batch.product_name.find(fangan.product_name) >= 0:
                flag_find = True
                data.append([wan_cheng_ri, batch.product_name, batch.batch, batch.yan_mo_hou, batch.jia_liao_hou,
                             batch.san_lei, batch.fan_hui_you, fangan.buliang_name, fangan.chuliliang])
                break
        if not flag_find:
            flag_find = False
            # 兑入同型号10%
            data.append([wan_cheng_ri, batch.product_name, batch.batch, batch.yan_mo_hou, batch.jia_liao_hou,
                         batch.san_lei, batch.fan_hui_you, '理论可处理同型号10%', 0.1])
    excel.set_range_value(1, 2, data)
    excel.save()


def jisuan_kunchun(out_file):
    excel = Excel(out_file)
    excel.select('不良库存')
    liushui = ProductLiushui.query.filter(and_(ProductLiushui.wan_cheng_ri >= datetime(2016, 5, 31),
                                               ProductLiushui.wan_cheng_ri <= datetime(2016, 6, 30))).all()
    buliangfang = BuliangFangan.query.all()
    kunchun = BuliangKuchun.query.all()
    data = []


def tongji_ruku(dest_filename='empty_book.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "品名"
    ws.cell(row=1, column=2).value = "批号"
    ws.cell(row=1, column=3).value = "配料日期"
    ws.cell(row=1, column=4).value = "生产完成日期"
    ws.cell(row=1, column=5).value = "入库日期"
    ws.cell(row=1, column=6).value = "入库品名"
    ws.cell(row=1, column=7).value = "生产打包量"
    ws.cell(row=1, column=8).value = "入库量"
    ws.cell(row=1, column=9).value = "是否正常"
    row = 2
    products = ProductLiushui.query.filter(and_(ProductLiushui.wan_cheng_ri >= datetime(2016, 7, 1),
                                                ProductLiushui.kind != "固化剂",
                                                ProductLiushui.kind != "色浆")).all()
    for product in products:
        is_ok = True
        product_count = product.gu_hua_ji + product.ru_ku_liang
        ruku_count = 0
        ruku_names = []
        ruku_riqis = []
        ruku = CangkuLiushui.query.filter(and_(CangkuLiushui.batch == product.batch,
                                               CangkuLiushui.yewu_type == "产品进仓")).all()
        for item in ruku:
            ruku_names.append(item.product_name)
            ruku_riqis.append(datetime.strftime(item.jilu_date, '%Y-%m-%d'))
            if item.amount:
                ruku_count += float(item.amount)
        if product_count != ruku_count:
            is_ok = False

        ws.cell(row=row, column=1).value = product.product_name
        ws.cell(row=row, column=2).value = product.batch
        ws.cell(row=row, column=3).value = datetime.strftime(
            product.product_date, '%Y-%m-%d') if product.product_date else None
        ws.cell(row=row, column=4).value = datetime.strftime(
            product.wan_cheng_ri, '%Y-%m-%d') if product.wan_cheng_ri else None
        ws.cell(row=row, column=5).value = ",".join(ruku_riqis)
        ws.cell(row=row, column=6).value = ",".join(ruku_names)
        ws.cell(row=row, column=7).value = product_count
        ws.cell(row=row, column=8).value = ruku_count
        ws.cell(row=row, column=9).value = is_ok

        row += 1

    wb.save(filename=dest_filename)


if __name__ == "__main__":
    # init_buliang_fangan()
    # load_buliang_kuchun(os.path.join(data_dir, '惠东不良品仓6月份流水帐2.xlsx'), start_row=4, stop_row=141)
    # load_cangku_liushui(os.path.join(data_dir, '7月份流水账.xlsx'), start_row=4, stop_row=4611)
    # load_product_liushui(os.path.join(data_dir, '7月份生产跟踪损耗日报表.xlsx'), start_row=15, stop_row=1509)
    # jisuan_ruku_liushui(os.path.join(data_dir, '6月理论处理量.xlsx'))
    # jisuan_product_liushui(os.path.join(data_dir, '6月理论处理量.xlsx'))
    # load_product_class(os.path.join(data_dir, '产品分类更新日期（20160721）-1.xls'))
    # chuli_liushuizhang(os.path.join(data_dir, '7月份流水账.xlsx'))
    tongji_ruku(os.path.join(data_dir, '生产量与入库量分析表.xlsx'))

    pass
