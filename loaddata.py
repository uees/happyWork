import argparse
import re
import os
import logging

import requests
from openpyxl import load_workbook

import settings

logging.getLogger("loaddata").setLevel(logging.WARNING)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
    datefmt='%a, %d %b %Y %H:%M:%S',
    filename='cataline.log',
    filemode='w'
)


def login(host, email, password):
    response = requests.post(f'{host}/api/auth/login', data={
        'email': email, 'password': password
    }, headers={
        'X-Requested-With': 'XMLHttpRequest'
    })

    body = response.json()

    return body.get('data').get('token')


def get_recyclable_type(product_name):
    recyclable_type = ''
    if product_name.find('SP') >= 0 and product_name.find('内袋') >= 0:
        recyclable_type = 'bucket'
    elif product_name.find('A-9060') >= 0 and product_name.find('内袋') >= 0:
        recyclable_type = 'bucket'
    elif product_name.find('内袋') >= 0:
        recyclable_type = 'box'

    return recyclable_type


def get_amount(weight, spec, recyclable_type):
    amount = 0
    match = re.match(r'^\d+\.?\d+', spec)
    if match:
        per_weight = float(match.group())
        if recyclable_type == 'box' and per_weight < 10:
            per_weight = 10

        if per_weight != 0:
            amount = int(weight / per_weight)

    return amount


def entering_warehouse(host, token, product_name, product_batch, spec, weight, entered_at, made_at, current_row):
    if product_name and product_batch and weight and made_at:
        print(f"Entering Warehouse 第{current_row}行：", entered_at, product_name, product_batch, spec)
        recyclable_type = get_recyclable_type(product_name)
        amount = get_amount(weight, spec, recyclable_type)
        response = requests.post(f'{host}/api/entering-warehouses', data={
            'product_name': product_name,
            'product_batch': product_batch,
            'spec': spec,
            'weight': abs(weight),
            'entered_at': entered_at,
            'made_at': made_at,
            'recyclable_type': recyclable_type,
            'amount': amount,
        }, headers={
            'X-Requested-With': 'XMLHttpRequest',
            'Authorization': f'Bearer {token}'
        })

        if response.status_code != 201:
            logging.error(f'第 {current_row} 行数据入库失败')


def shipment(host, token, custmor, product_name, product_batch, spec, weight, created_at, current_row):
    if product_name and weight and custmor:
        print(f"Shipment 第{current_row}行：", created_at, product_name, product_batch, spec)
        recyclable_type = get_recyclable_type(product_name)
        amount = get_amount(weight, spec, recyclable_type)
        data = {
            'customer_id': custmor,
            'product_name': product_name,
            'product_batch': product_batch,
            'spec': spec.upper() if isinstance(spec, str) else '',
            'weight': abs(weight),
            'created_at': created_at,
            'recyclable_type': recyclable_type,
            'amount': amount,
        }
        response = requests.post(f'{host}/api/shipments', data=data, headers={
            'X-Requested-With': 'XMLHttpRequest',
            'Authorization': f'Bearer {token}'
        })

        if response.status_code != 201:
            logging.error(f'第 {current_row} 行数据发货失败')


def load_file(filename, start_row=3):
    host = os.getenv('API_HOST')
    email = os.getenv('ADMIN_USER')
    passwd = os.getenv('ADMIN_PASS')
    token = login(host, email, passwd)

    wb = load_workbook(filename)
    ws = wb['成品流水账']
    current_row = start_row
    for row in ws[f'A{start_row}:J{ws.max_row}']:
        _type, date, NO, custmor, code, name, spec, batch, weight, made_at = row

        # 跳过不是内袋装的油
        if not name.value or name.value.find('内袋') == -1:
            continue

        if _type.value == "产品进仓":
            entering_warehouse(
                host=host,
                token=token,
                product_name=name.value,
                product_batch=batch.value,
                spec=spec.value,
                weight=weight.value,
                entered_at=date.value,
                made_at=made_at.value,
                current_row=current_row
            )
        elif _type.value == "往来销售":
            if not custmor.value:
                continue

            shipment(
                host=host,
                token=token,
                custmor=custmor.value,
                product_name=name.value,
                product_batch=batch.value,
                spec=spec.value,
                weight=weight.value,
                created_at=date.value,
                current_row=current_row
            )

        current_row += 1


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='导入流水数据')
    parser.add_argument("excel_file", help="文件名")
    parser.add_argument("-i", "--index", type=int, help="excel中需要生成报告的起始行")

    args = parser.parse_args()

    if args.index:
        start_row = args.index
    else:
        start_row = 3

    load_file(args.excel_file, start_row)
