# -*- coding: utf-8 -*-
"""
Created on 2015年7月10日

@author: Wan
"""
from openpyxl import load_workbook
from sqlalchemy import and_

from _functools import reduce
from database import IQCMaterial, Product, db_session


def search_product(keywords):
    criteria = []
    for keyword in keywords.split():
        criteria.append(Product.internal_name.ilike('%{}%'.format(keyword)))
    q = reduce(and_, criteria)
    return Product.query.filter(q)


def get_product_by_id(pid):
    return Product.query.filter_by(id=pid).first()


def get_product_by_internal_name(internal_name):
    return Product.query.filter_by(internal_name=internal_name).first()


def insert_product(product):
    db_session.add(product)
    db_session.commit()


def insert_product_to_xlsx(product, filename, sheet):
    wb = load_workbook(filename)
    ws = wb[sheet]
    index = ws.max_row + 1
    ws["A{}".format(index)] = product.internal_name
    ws["B{}".format(index)] = product.template
    ws["C{}".format(index)] = product.viscosity
    ws["D{}".format(index)] = product.viscosity_width
    ws["E{}".format(index)] = product.market_name
    ws["J{}".format(index)] = product.color
    wb.save(filename)


def init_product_data(file, sheet):
    """ 从 database.xlsx 读取数据 """
    wb = load_workbook(filename=file)
    ws = wb[sheet]
    for row in ws['A2:J{}'.format(ws.max_row)]:
        (internal_name, template, viscosity, viscosity_width, market_name, category,
         part_a, part_b, ratio, color) = [cell.value for cell in row]
        if internal_name:
            product = Product(internal_name=internal_name,
                              template=template,
                              viscosity=viscosity,
                              viscosity_width=viscosity_width,
                              market_name=market_name,
                              category=category,
                              part_a=part_a,
                              part_b=part_b,
                              ratio=ratio,
                              color=color)
            db_session.add(product)
    db_session.commit()
    print("插入了 %s 行数据到data/database.sdb3." % str(ws.max_row - 1))


def init_materials(filename, sheet):
    wb = load_workbook(filename)
    ws = wb[sheet]
    for row in ws['A2:D{}'.format(ws.max_row)]:
        material_name, qc_items, spec, qc_method = [cell.value for cell in row]
        if not material_name:
            continue
        qc_items = qc_items.replace(',', '、').replace('，', '、')
        material = IQCMaterial(name=material_name,
                               qc_items=qc_items,
                               spec=spec,
                               qc_method=qc_method)
        db_session.add(material)
    db_session.commit()
    print("插入了 %s行数据到data/database.sdb3." % str(ws.max_row - 1))
