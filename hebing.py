#!/usr/bin/env python
# -- coding: utf-8 -*-
'''
Created on 2016年8月20日

@author: Wan
'''
import os
from openpyxl import load_workbook
from common import module_path
from database import Hebing, reset_table, db_session


data_dir = os.path.join(module_path(), 'data/1')
to_filename = os.path.join(data_dir, 'bbb.xlsx')
from_filename = os.path.join(data_dir, '2012-2016年供应商列表.xlsx')

to_wb = load_workbook(to_filename)
to_ws = to_wb.get_sheet_by_name('Sheet1')


def load_fenlei(filename, init=False):
    wb = load_workbook(filename, data_only=True)
    ws = wb.get_sheet_by_name('1')
    fenlei = []
    for row in ws.iter_rows('A2:B{}'.format(ws.max_row)):
        fenlei.append([cell.value for cell in row])
        if init:
            customer, code = [cell.value for cell in row]
            db_session.add(Hebing(code=code, customer=customer))
    if init:
        db_session.commit()
    return fenlei


def hebing(from_, to_):
    from_wb = load_workbook(from_filename, data_only=True)


def write_customer():
    fenlei = load_fenlei(os.path.join(data_dir, '供应商编码1.xlsx'))
    row = 2
    for customer, code in fenlei:
        to_ws.cell(row=row, column=1).value = code
        to_ws.cell(row=row, column=2).value = customer
        row += 1
    to_wb.save(to_filename)


def load_2012_caigou():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2012年')
    _ws = to_wb.create_sheet('cg_2012')
    i = 2
    for row in range(2, 109):
        customer, value = [ws.cell(row=row, column=2).value, ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.cg_2012 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2013_caigou():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2013年')
    _ws = to_wb.create_sheet('cg_2013')
    i = 2
    for row in range(2, 118):
        customer, value = [ws.cell(row=row, column=3).value,
                           ws.cell(row=row, column=5).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.cg_2013 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2014_caigou():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2014年')
    _ws = to_wb.create_sheet('cg_2014')
    i = 2
    for row in range(2, 98):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=4).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.cg_2014 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2015_caigou():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2015年')
    _ws = to_wb.create_sheet('cg_2015')
    i = 2
    for row in range(2, 98):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.cg_2015 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2012_yingfu():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2012年应付')
    _ws = to_wb.create_sheet('yf_2012')
    i = 2
    for row in range(2, 108):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=4).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.yf_2012 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2013_yingfu():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2013应付')
    _ws = to_wb.create_sheet('yf_2013')
    i = 2
    for row in range(2, 99):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=4).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.yf_2013 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2014_yingfu():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2014年应付')
    _ws = to_wb.create_sheet('yf_2014')
    i = 2
    for row in range(2, 135):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=4).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.yf_2014 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2015_yingfu():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2015年应付')
    _ws = to_wb.create_sheet('yf_2015')
    i = 2
    for row in range(2, 167):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=5).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.yf_2015 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2012_yue():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2012余额')
    _ws = to_wb.create_sheet('ye_2012')
    i = 2
    for row in range(4, 110):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.ye_2012 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2013_yue():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2013余额')
    _ws = to_wb.create_sheet('ye_2013')
    i = 2
    for row in range(4, 147):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.ye_2013 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2014_yue():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2014余额')
    _ws = to_wb.create_sheet('ye_2014')
    i = 2
    for row in range(4, 150):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.ye_2014 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


def load_2015_yue():
    from_wb = load_workbook(from_filename, data_only=True)
    ws = from_wb.get_sheet_by_name('2015余额')
    _ws = to_wb.create_sheet('ye_2015')
    i = 2
    for row in range(4, 137):
        customer, value = [ws.cell(row=row, column=2).value,
                           ws.cell(row=row, column=3).value]
        record = Hebing.query.filter(Hebing.customer == customer).first()
        if record:
            record.ye_2015 = value
            db_session.commit()
        else:
            _ws.cell(row=i, column=1).value = customer
            _ws.cell(row=i, column=2).value = value
            i += 1
    to_wb.save(to_filename)


if __name__ == "__main__":
    #load_fenlei(os.path.join(data_dir, '供应商编码1.xlsx'), True)
    # load_2012_caigou()
    # load_2012_yingfu()
    # load_2012_yue()
    # load_2013_caigou()
    # load_2013_yingfu()
    # load_2013_yue()
    # load_2014_caigou()
    # load_2014_yingfu()
    # load_2014_yue()
    # load_2015_caigou()
    # load_2015_yingfu()
    load_2015_yue()
