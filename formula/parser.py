# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime

from openpyxl import load_workbook

from settings import formula as formula_template


class FormulaParser(object):
    """ 分析一个excel文件，获取产品配方信息"""

    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = load_workbook(self.filepath, data_only=True)
        self.worksheets_num = len(self.workbook.worksheets)
        self.formulas = []

    def parse(self):
        """ return all formulas in self.workbook
        formula is a dict(
            name<str>,
            category<str>,
            version<str>,
            common_name<str>,
            description<str>,
            date<datetime>,
            materials<list>,
            extend_materials<list>,
            metas<list>
        )
        @see settings.formula
        """
        formulas = []
        formula = self.parse_filename()
        materials = self.get_mixing_materials()
        products = self.get_products()
        for product_name, sheet_name in products:
            formula['name'] = product_name  # 重置为车间使用的产品名
            extends_info = self.get_extends_info(sheet_name)
            materials.extend(extends_info['materials'])
            formula['materials'] = materials
            formula['jialiao_yaoqiu'] = extends_info['yaoqiu']
            formulas.append(formula)
        return formulas

    def parse_filename(self):
        """
        解析文件名
        return dict(name, created_at, description, version)
        """
        filename = os.path.basename(self.filepath).replace("（", "(").replace("）", ")")
        pattern = re.compile(r'''^(?P<created_at>\d{4}-\d{1,2}-\d{1,2})?
                                  (?P<name>.+)
                                  \(
                                  (?P<version>B-\d{1,2})
                                  \)
                                  (?P<description>.*)
                                  \.xlsx$''', re.X)
        match = pattern.match(filename)
        if match:
            formula = match.groupdict()
            formula['name'] = formula['name'].strip().strip("_")
        else:
            formula = dict(name=os.path.splitext(filename)[0],
                           created_at=datetime.utcnow().strftime("%Y-%m-%d"),
                           version='B-01',
                           description='')

        return formula_template.copy().update(formula)

    def get_products(self):
        """ return a [(product_name, sheet_name)] list """
        products = []
        sheets = self.workbook.get_sheet_names()
        for sheet in sheets:
            if sheet.find('配料单') < 0:  # 跟踪单
                name = self.workbook.get_sheet_by_name(sheet).cell("B4").value
                if name:
                    products.append((name, sheet))
                else:
                    raise Exception(f'在 {self.filepath} 的 sheet {sheet} 中没有找到产品名, '
                                    '请检查是否为不标准的配方格式。')
        return products

    def get_mixing_materials(self, start=8):
        """
        获取配料表
        return a list[(name, amount, location)]
        """
        materials = []
        try:
            ws = self.workbook.get_sheet_by_name('配料单')
        except:
            ws = self.workbook.worksheets[0]

        for row in ws.iter_rows("B8:C{}".format(ws.max_row)):
            name, amount = [cell.value for cell in row]
            if name and isinstance(amount, float) or isinstance(amount, int):
                materials.append((name, amount, '配料'))

        return materials

    def get_extends_info(self, sheet):
        """
        获取加料信息
        return a dict(materials<list>, requirements<list>)
        """
        info = dict(materials=[], requirements=[])
        ws = self.workbook.get_sheet_by_name(sheet)
        title = ws.cell("A2").value
        if title is None or title.replace(" ", "") == u"RoHS配料生产记录表":  # 配料单
            return info

        start_row = self.get_extends_start_row(ws)

        for row in range(start_row, ws.max_row):
            name = ws.cell(row=row, column=1).value
            amount = ws.cell(row=row, column=3).value
            if name == '返回油墨':  # 加料信息结束行
                break

            if name:
                if isinstance(amount, float) or isinstance(amount, int):
                    info['materials'].append(dict(name=name, amount=amount, unit='%', workshop='加料'))
                elif amount == '稀释剂':
                    info['materials'].append(dict(name=name, amount=0, unit='kg', workshop='加料'))
                elif not amount:
                    info['requirements'].append(name)

        return info

    def get_extends_start_row(self, ws):
        """获取加料信息起始行"""
        for row in range(13, ws.max_row):
            if ws.cell(row=row, column=1).value == "物料名称":
                return row
