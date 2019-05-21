# -*- coding: utf-8 -*-
"""
Created on 2016年6月15日

@author: Wan
"""
import os
import re
from datetime import datetime

from openpyxl import load_workbook


class FormulaAnalyst(object):
    """ 分析一个excel文件，获取产品配方信息"""

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename=self.filename, data_only=True)
        self.ws_count = len(self.wb.worksheets)

    def get_formulas(self):
        """ return all formulas in self.filename
        formula is dict(name<str>, version<str>, info<str>, date<datetime>,
        materials<list>, jialiao_yaoqiu<list>)"""
        formulas = []
        formula = self.match_filemame()
        materials = self.get_mixing_materials()
        products = self.get_product_name_in_sheet()
        for product_name, sheet_name in products:
            formula['name'] = product_name  # 重置为车间使用的产品名
            jialiao_info = self.get_jiaoliao_info(sheet_name)
            materials.extend(jialiao_info['materials'])
            formula['materials'] = materials
            formula['jialiao_yaoqiu'] = jialiao_info['yaoqiu']
            formulas.append(formula)
        return formulas

    def match_filemame(self):
        """ return dict(name, date, info, version) """
        filename = self.filename.replace("（", "(").replace("）", ")")
        pattern = re.compile(r'''^(?P<date>\d{4}-\d{1,2}-\d{1,2})?
                                  (?P<name>.+)
                                  \(
                                  (?P<version>B-\d{1,2})
                                  \)
                                  (?P<info>.*)
                                  \.xlsx$''', re.X)
        match = pattern.match(filename)
        if match:
            formula = match.groupdict()
            formula['name'] = formula['name'].strip().strip("_")
            return formula

        return dict(name=os.path.splitext(filename)[0],
                    date=datetime.utcnow(),
                    version='B-01',
                    info='')

    def get_product_name_in_sheet(self):
        """ return a [(product_name, sheet_name)] list """
        products = []
        sheets = self.wb.get_sheet_names()
        for sheet in sheets:
            if sheet.find('配料单') < 0:  # 跟踪单
                name = self.wb.get_sheet_by_name(sheet).cell("B4").value
                if name:
                    products.append((name, sheet))
                else:
                    raise Exception('在 %s 的sheet %s 中没有找到产品名, 请检查是否为不标准的配方格式。' %
                                    (self.filename, sheet))

        assert len(products) == len(sheets) - 1
        return products

    def get_mixing_materials(self, start=8):
        """ return a list[(name, amount, location)] """
        materials = list()
        try:
            ws = self.wb.get_sheet_by_name('配料单')
        except:
            ws = self.wb.worksheets[0]

        for row in ws.iter_rows("B8:C{}".format(ws.max_row)):
            name, amount = [cell.value for cell in row]
            if name and isinstance(amount, float) or isinstance(amount, int):
                materials.append((name, amount, '配料'))

        return materials

    def get_jiaoliao_info(self, sheet):
        """ return a dict(materials<list>, yaoqiu<list>) """
        info = dict(materials=[], yaoqiu=[])
        ws = self.wb.get_sheet_by_name(sheet)
        title = ws.cell("A2").value
        if title is None or title.replace(" ", "") == u"RoHS配料生产记录表":  # 配料单
            return info

        row = 13
        while ws.cell(row=row, column=1).value != "物料名称":  # 获取加料信息起始行
            row += 1
            if row > ws.max_row:  # 防止死循环
                break

        while True:  # 获取加料信息结束行
            row += 1
            name = ws.cell(row=row, column=1).value
            amount = ws.cell(row=row, column=3).value
            if name == '返回油墨' or row > ws.max_row:  # 加料信息结束行
                break
            if name:
                if isinstance(amount, float) or isinstance(amount, int):
                    info['materials'].append((name, amount, '加料'))
                elif amount == '稀释剂':
                    info['materials'].append((name, 0, '加料'))
                elif not amount:
                    info['yaoqiu'].append(name)
        return info
