import os
import re
import glob
from openpyxl import Workbook

from formula.parser import FormulaParser
from settings import BASE_DIR
from common import load_config

CONFIG = load_config()
WORK_ORDERS_PATH = os.path.join(BASE_DIR, CONFIG.get('default', 'formula_dir'))


def extract_viscosity():
    """提取粘度数据"""
    all_formula_files = glob.glob(f"{WORK_ORDERS_PATH}/**/*.xlsx", recursive=True)

    result = []
    for filepath in all_formula_files:
        parser = FormulaParser(filepath)
        formulas = parser.parse()
        for formula in formulas:
            after_adding_requirement = formula['metas']['after_adding_requirement']
            if isinstance(after_adding_requirement, list):
                for requirement in after_adding_requirement:
                    if requirement.find("粘度要求") >= 0:
                        result.append(dict(name=formula['name'], viscosity=requirement))
            elif after_adding_requirement and isinstance(after_adding_requirement, str):
                if after_adding_requirement.find("粘度要求") >= 0:
                    result.append(dict(name=formula['name'], viscosity=after_adding_requirement))

    return _extract_value(result)


def _extract_value(data):
    # .+? 非贪婪模式  .+ 贪婪模式
    pattern1 = re.compile(r'.+?(\d+)\s*±\s*(\d+)\s*d[P|p]a.*')
    pattern2 = re.compile(r'.+?(\d+)\s*~\s*(\d+)\s*d[P|p]a.*')

    result = []
    for item in data:
        viscosity = item['viscosity'].replace('～', '~').replace('-', '~')
        match = pattern1.match(viscosity)
        if match:
            middle, extent = match.groups()
            result.append({
                "name": item["name"],
                "min": int(middle) - int(extent),
                "max": int(middle) + int(extent),
            })
            continue

        match = pattern2.match(viscosity)
        if match:
            _min, _max = match.groups()
            result.append({
                "name": item["name"],
                "min": int(_min),
                "max": int(_max),
            })
            continue

        result.append(item)

    return result


def viscosity2excel(data):
    wb = Workbook()
    ws = wb.active

    for index, item in enumerate(data):
        ws.cell(row=index + 1, column=1, value=item['name'])
        if 'min' in item:
            ws.cell(row=index + 1, column=2, value=item['min'])
            ws.cell(row=index + 1, column=3, value=item['max'])
        if 'viscosity' in item:
            ws.cell(row=index + 1, column=4, value=item['viscosity'])

    wb.save(filename=os.path.join(BASE_DIR, 'data/viscosity.xlsx'))


def viscosity2db(data, host, user, password, dbname):
    import pymysql
    import json

    # 打开数据库连接
    db = pymysql.connect(host, user, password, dbname)

    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = db.cursor()

    # 使用 execute()  方法执行 SQL 查询
    cursor.execute("SELECT VERSION()")

    # 使用 fetchone() 方法获取单条数据.
    data = cursor.fetchone()

    print("Database version : %s " % data)

    for item in data:
        if 'min' in item:
            meta = {
                "spec_viscosity": f"{data['min'] - data['max']}"
            }
            # todo to database

    # 关闭数据库连接
    db.close()


def extract_fineness():
    """提取细度数据"""
    pass
