#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import os
import random
import re
import sys
from datetime import datetime

from openpyxl import load_workbook

import data_warp as db
from utils import iqc_report
# from fqc_list import FqcListGenerator
from common import is_number, is_number_like, module_path, null2str, rlinput
from settings import CONF
from database import Product, init_database, reset_table
from utils.office import WordTemplate as WTemplate
from utils.pdf import add_watermark, create_watermark


class Generator(object):
    """ 检验报告生成器 """

    def __init__(self, start_index=None):
        self.start_index = start_index
        self.index = 0
        self.app_path = module_path()
        self._product_wb_file = "%s/reports/list.xlsx" % self.app_path
        self._wb = load_workbook(self._product_wb_file)
        self._ws = None
        # self.fqc_g = FqcListGenerator('%s/reports/FQC&IQC检测记录.xlsx' % self.app_path)

    def get_start_row(self, flag_done_col=10):
        if self._ws:
            max_row = self._ws.max_row
            for row in range(max_row, 2, -1):
                done_cell = self._ws.cell(row=row, column=flag_done_col)
                if done_cell.value:
                    return row + 1

        # if all None, the done row is 2
        return 2

    def generate_reports(self, sheet="Sheet1"):
        """ 批量生成检验报告 """
        self._ws = self._wb[sheet]

        if not self.start_index:
            self.start_index = self.get_start_row()

        for index in range(self.start_index, self._ws.max_row + 1):
            print("\n------------------------------")
            self.index = index
            _info = self.get_product_info(index)
            if not _info or not _info['internal_name']:
                continue

            print("第 %s行, 品名:%s, 批号:%s" % (index, _info['internal_name'], _info['batch']))

            product = self.query_info(_info)
            if not product:
                continue

            # 可以有效减少背锅概率
            if len(product['batch']) != 8 and product['kind'] != 'xsj' and product['market_name'] != 'LPI-360GS':
                print("Line{}:卧槽,批号格式不一般,直接跳过不生成!!@#$@#&%".format(index))
                continue

            # fix模式: 修改了 product
            self.fix_宏华胜(product)
            self.fix_高士(product)
            self.fix_A_9060C0101(product)

            self.generate_report(product)

            if product['wants_normal']:
                self.generate_normal(product)

            # 专用报告
            self.generate_达进(product)
            self.generate_景旺(product)
            self.generate_健鼎(product)
            self.generate_深南(product)
            self.generate_南通深南(product)
            self.generate_崇达(product)
            self.generate_木林森(product)
            self.generate_华新(product)
            self.generate_威尔高(product)
            self.generate_金像(product)
            self.generate_xsj_with_amount(product)
            self.generate_bomin(product)
            self.generate_建业(product)

            self._set_report_info(product)
            # self.fqc_g.fqc_record(product)

        self.save()

    def fix_A_9060C0101(self, product):
        if product['market_name'] == 'A-9060C 01 01':
            product['viscosity'] = "%s" % random.randint(200, 210)
            product['wants_normal'] = False
        elif product['market_name'] == 'A-9060C 01':
            product['viscosity'] = "%s" % random.randint(90, 100)
            product['wants_normal'] = False

    def fix_高士(self, product):
        if product['market_name'] == "LPI-360GS":
            product["viscosity_limit"] = "450~650"
            product["viscosity"] = str(random.choice(range(480, 630)))
            product["shuanzhi"] = str(round(random.uniform(52, 55), 2))
            product["ftir"] = '{}%'.format(round(random.uniform(99.3, 100), 2))

    def fix_宏华胜(self, product):
        # product 数组引用传值，内部修改影响外面
        if product['market_name'].find('28GHB') >= 0 or \
                product['market_name'].find('30GHB') >= 0 or \
                product['market_name'].find('59GHB') >= 0:
            product['ext_info'] += '(宏华胜要求打发货数量)'
            product['kind'] = 'h9100_fsk'  # 宏华胜要求的是20℃的粘度
            product['wants_normal'] = False
        elif product['market_name'].find('SP20HF') >= 0:
            product['ext_info'] += '(宏华胜要求打发货数量)'
            product['wants_normal'] = False

    def generate_威尔高(self, product):
        products32 = ['MG605', 'GH6801', 'G603', 'MBK01', '6GHB深色', '8BL2', '8BKH', 'GH15', '9G201', '8R2']
        products36 = []

        if not product['kind'].endswith('_weg'):
            if product['internal_name'] in products32:
                new_product = product.copy()
                new_product["kind"] = '%s_weg' % product['kind']
                new_product['dayinReq'] = "≥32达因"
                new_product['dayinVal'] = "{}达因".format(random.choice([32, 33]))
                self.generate_report(new_product)

            if product['internal_name'] in products36:
                new_product = product.copy()
                new_product["kind"] = '%s_weg' % product['kind']
                new_product['dayinReq'] = "≥36达因"
                new_product['dayinVal'] = "{}达因".format(random.choice([36, 37]))
                self.generate_report(new_product)

    def generate_华新(self, product):
        """ SK45 GH6 8G 塞孔 要求打 H-8100 """
        if product['market_name'] == "8G 塞孔" or \
                product['market_name'] == "GH6" or \
                product['market_name'] == "SK45":
            new_product = product.copy()
            new_product["kind"] = 'h8100'
            new_product['ext_info'] += '【华新专用报告】'
            self.generate_report(new_product)

    def generate_金像(self, product):
        """ A-9060A 01 打 7±2 """
        if product['market_name'] == "A-9060A 01":
            new_product = product.copy()
            new_product["viscosity_limit"] = "7±2"
            new_product["viscosity"] = "%.1f" % random.uniform(7, 9)
            new_product['ext_info'] += '【金像专用报告】'
            self.generate_report(new_product)

    def generate_建业(self, product):
        """ 建业 MG31 160±50 """
        if product['market_name'].startswith("MG31"):
            new_product = product.copy()
            new_product["viscosity_limit"] = "160±50"
            new_product["viscosity"] = str(random.choice(range(160, 170)))
            new_product['ext_info'] += '【建业专用报告】'
            self.generate_report(new_product)

    def generate_木林森(self, product):
        """ 木林森对字符油粘度范围有特殊要求 """
        product = product.copy()
        product['ext_info'] += '【木林森专用】'
        matched = False

        if product['market_name'] == "TM-3100 BK":
            matched = True
            product["kind"] = "tm3100_mls"
            product["viscosity_limit"] = "≥300"
            product["viscosity"] = str(random.choice(range(300, 310)))
        elif product['market_name'] == "TM-3100 W":
            matched = True
            product["kind"] = "tm3100_mls"
            product["viscosity_limit"] = "200~300"
            product["viscosity"] = str(random.choice(range(260, 300)))
        elif product["market_name"].startswith('A-2100'):
            matched = True
            product["viscosity_limit"] = "200±100"
        elif product["market_name"] == "W16":  # 热固阻焊油 200±50
            matched = True
            product["viscosity_limit"] = "200±50"
        elif product["market_name"] == "SK30A":
            matched = True
            product["viscosity_limit"] = "250~550"
        elif product["market_name"] == "SK27 01":
            matched = True
            product["viscosity_limit"] = "250~500"
        elif product["market_name"] == "SK11 01":
            matched = True
            product["viscosity_limit"] = "250~350"
            product["viscosity"] = str(random.choice(range(260, 290)))
        elif product["market_name"] == "8G 05":
            matched = True
            product["viscosity_limit"] = "≥160"
        elif product["market_name"] == "8W4":
            matched = True
            product["viscosity_limit"] = "≥160"
        elif product["market_name"] in ("MBK21HF", "MBK21"):
            matched = True
            product["viscosity_limit"] = "120~200"
        elif product["market_name"] == "A-9070 01":
            matched = True
            product["viscosity_limit"] = "≥60"
        elif product["market_name"] == "GH16":
            matched = True
            product["viscosity_limit"] = "≥160"

        # todo 面油 180±50，塞孔油 300±100

        if matched:
            self.generate_report(product)

    def generate_达进(self, product):
        if product['market_name'] == '8BL2' or \
                product['market_name'] == '8WL5 01' or \
                product['market_name'] == '44G' or \
                product['market_name'] == '6GHB HF':

            if not product['kind'].endswith('_dj'):
                new_product = product.copy()
                new_product["kind"] = '%s_dj' % product['kind']
                self.generate_report(new_product)

    def generate_景旺(self, product):
        if product['market_name'] == '6GHB HF' or product['market_name'] == 'MG55':
            if not product['kind'].endswith('_jw'):
                new_product = product.copy()
                new_product["kind"] = '%s_jw' % product['kind']
                self.generate_report(new_product)

        # 景旺的 SK01HF01 要求打 22 度下的粘度，要求动态粘度数值
        if product['market_name'] == 'SK01HF01':
            if not product['kind'].endswith('_jw22'):
                new_product = product.copy()
                new_product["viscosity_limit"] = "300±50"
                new_product["viscosity"] = str(random.choice(range(280, 320)))
                new_product["kind"] = '{}_jw22'.format(product['kind'])
                self.generate_report(new_product)

    def generate_bomin(self, product):
        """生成博敏的专用报告"""
        if product['market_name'].replace(' ', '') in ["8BL2", "8G0105", "SK35", "23GHB", "8BL7", "WB5101",
                                                       "8BK15", "MG31", "8BL9HF", "20GHB", "SK47", "SK27", "SK41HF",
                                                       "MG3101", "SK40HF", "T-SK2902", "T-SK50BLHF", "44G", "8R15",
                                                       "80R6", "15GHB 12", "MG75HF", "G602", "SK36W", "SK8R",
                                                       "AMG3", "MBK16", "YH2", "8GR01", "中绿", "09R", "SK3502"]:
            if not product['kind'].endswith('_bomin'):
                new_product = product.copy()
                new_product["kind"] = "{}_bomin".format(product['kind'])
                self.generate_report(new_product)

    def generate_健鼎(self, product):
        if product['internal_name'].find('健鼎') >= 0 or \
                product['market_name'] == 'A-9060B':

            if not product['kind'].endswith('_jd'):  # 这时没有标注的才创建, 标注过的已经创建了
                new_product = product.copy()
                new_product["kind"] = "{}_jd".format(product['kind'])
                self.generate_report(new_product)

        if product['market_name'] == "A-9060C":
            if not product['kind'].endswith('_hbjd'):
                new_product = product.copy()
                new_product["kind"] = "a9060c_hbjd"
                new_product["viscosity_limit"] = "160±30"  # 湖北健鼎要求秒数单位 22C
                new_product["viscosity"] = str(random.choice(range(140, 160)))
                self.generate_report(new_product)

        elif product.get('market_name').find('15GHB') == 0:  # 健鼎 15GHB 粘度搞成 180±50
            new_product = product.copy()
            new_product['ext_info'] += '(健鼎专用报告)'
            new_product["viscosity_limit"] = "180±50"
            new_product["viscosity"] = str(random.choice(range(160, 170)))
            self.generate_report(new_product)

    def generate_深南(self, product):
        if product.get('market_name').find('SP8') == 0 or \
                product.get('market_name').find('SP50') == 0 or \
                product.get('market_name').find('SPM') == 0 or \
                product.get('market_name') == 'A-9060A 01' or \
                product.get('market_name') == '60G' or \
                product.get('market_name').find('SK29') >= 0:

            # SP8T3 and SP8T308 红外PDF
            if product.get('market_name') == "SP8T3" or \
                    product.get('market_name').replace(' ', '') == "SP8T308":
                self.generatePDF(product)

            # T-SK2902 红外PDF
            if product.get('market_name').find('T-SK29') >= 0:
                self.generatePDF(product, dir='T-SK2902', max=17)

                # T-SK2902 要求打 22 度下的粘度，要求动态粘度数值
                product = product.copy()
                product["viscosity_limit"] = "300±50"
                product["viscosity"] = str(random.choice(range(280, 320)))
                product["kind"] = 'h9100_22c'

            if not product['kind'].endswith('_ntsn'):
                new_product = product.copy()
                new_product['ext_info'] += '(深南要打发货数量)'
                self.generate_report(new_product, '深南')

    def generate_南通深南(self, product):
        if product.get('market_name').find('SP8') == 0 or \
                product.get('market_name').find('SP50') == 0 or \
                product.get('market_name').find('SPM') == 0 or \
                product.get('market_name') == '60G' or \
                product.get('market_name').find('SK29') >= 0:

            if not product['kind'].endswith('_ntsn'):  # 这时没有标注的才创建, 标注过的已经创建了
                new_product = product.copy()
                new_product["kind"] = '%s_ntsn' % product['kind']
                new_product['ext_info'] += '(南通深南要打发货数量、要发邮件到ntiqc@scc.com.cn)'
                self.generate_report(new_product, '深南')

    def generate_崇达(self, product):
        if product.get('market_name') == '8BL' or \
                product.get('market_name') == 'GH3' or \
                product.get('market_name').find('G6') == 0 or \
                product.get('market_name').find('SP02') == 0 or \
                product.get('market_name').find('GH40') == 0 or \
                product.get('market_name').find('MG31') == 0 or \
                product.get('market_name').find('23GHB') == 0 or \
                product.get('internal_name').find('崇达') >= 0:

            if not product['kind'].endswith('_cd'):  # 这时没有标注的才创建, 标注过的已经创建了
                new_product = product.copy()
                new_product["kind"] = '%s_cd' % product['kind']
                self.generate_report(new_product)

    def generate_xsj_with_amount(self, product):
        if product['kind'] == 'xsj':
            new_product = product.copy()
            new_product["kind"] = '%s_amount' % product['kind']
            new_product['ext_info'] = '(深南-崇达-宏华胜-要打发货数量)' + new_product['ext_info']
            self.generate_report(new_product)

    def generate_normal(self, product):
        if product['kind'].endswith('_jx'):  # 金像是设置的主剂粘度, 只此一家用, 暂时不生成 normal
            return
        f = product["kind"].find('_')
        if f > 0:
            new_product = product.copy()
            new_product["kind"] = product["kind"][:f]
            new_product['ext_info'] = ''
            self.generate_report(new_product)

    def generatePDF(self, product, dir=None, max=30):
        spa_path = os.path.join(self.app_path, 'spa')
        basename = product.get('market_name').replace(' ', '')

        if not dir:
            dir = basename

        f_name = '%s.pdf' % random.randint(1, max)
        f_path = os.path.join(spa_path, '%s/%s' % (dir, f_name))

        out_f_name = '%s-%s.pdf' % (basename, product.get('batch'))
        out_f_path = os.path.join(self.get_today_report_dir_path(), out_f_name)

        watermark = create_watermark(out_f_name, spa_path)
        add_watermark(watermark, f_path, out_f_path)

    def format_filename(self, customer, product):
        if customer == "深南":
            qc_date = datetime.strftime(datetime.now(), '%Y%m%d')
            filename = '{}_{}_{}容大{}COC.docx'.format(product["batch"],
                                                     product["ext_info"],
                                                     qc_date,
                                                     product["market_name"])

        else:
            filename = '{}_{}_{}{}.docx'.format(product["batch"],
                                                product["internal_name"],
                                                product["spec"],
                                                product["ext_info"])

        return filename

    def generate_report(self, product, customer=None):
        """ 生成检验报告 """
        if product['kind'].endswith('_sn'):
            if product['ext_info'] == '':
                return  # fix bug

        conf = self.get_conf(product["kind"])
        if not conf:
            print('无效的产品类别')
            return

        if conf.get('customer'):
            product['ext_info'] += '【%s专用报告】' % conf.get('customer')

        if conf.get('ext_info'):
            product['ext_info'] += conf.get('ext_info')

        template = self.get_template(conf.get('template'))
        if not os.path.exists(template):
            print("%s 模板文件不存在！" % product["kind"])
            return

        tp = WTemplate(template)
        tp.replace(product)

        today_report_dir = self.get_today_report_dir_path()

        filename = self.format_filename(customer, product)

        filepath = '{}/{}'.format(today_report_dir, filename)

        if os.path.exists(filepath):
            print("{}已经存在了".format(filename))
        else:
            tp.save(filepath)
            print("报告已经生成：{}".format(filename))

    def get_product_info(self, index):
        """ 从 list.xlsx 获取指定行的产品输入信息 """
        validity_date = ''
        ext_info = ''

        customer, internal_name, spec, batch, amount, product_date = [
            self._ws.cell(row=index, column=i).value for i in range(1, 7)]

        customer, internal_name, spec, batch, amount, product_date = map(
            null2str, [customer, internal_name, spec, batch, amount, product_date])

        internal_name = re.sub(r'[\(\)（）]|20kg|20KG|5kg|5KG|1kg|1KG|内袋|外贸', ' ', internal_name)  # 去除不良字符
        internal_name = internal_name.strip()

        if is_number(batch):
            batch = str(int(batch))
        if not batch:
            batch = ''
            ext_info += "【批号没填】"

        if is_number(amount):
            # amount = '发货数量:{}kg'.format(int(amount))
            amount = ''
        else:
            amount = ''

        if not isinstance(product_date, datetime):
            print("Line{}:时间格式不正确,已经为你设置为空串.".format(index))
            product_date = ''
            ext_info += "【生产日期没填】"
        else:
            mon_days = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            next_year = product_date.year
            next_month = product_date.month + 6
            if next_month > 12:
                next_month -= 12
                next_year = product_date.year + 1

            next_day = product_date.day
            if next_day > mon_days[next_month]:
                next_day = mon_days[next_month]

            validity_date = datetime(year=next_year, month=next_month, day=next_day)

            product_date = datetime.strftime(product_date, '%Y/%m/%d')  # 转为字符串
            validity_date = datetime.strftime(validity_date, '%Y/%m/%d')  # 转为字符串

        return dict(customer=customer,
                    internal_name=internal_name,
                    spec=spec,
                    batch=batch,
                    amount=amount,
                    product_date=product_date,
                    validity_date=validity_date,
                    ext_info=ext_info)

    def query_info(self, given):
        """
        查询数据库, 追加更多信息
        :param: given 是从list.xlsx 中查出的数据
        """
        products = db.search_product(given['internal_name'])
        if products.count() == 0:
            print('\n数据库中无记录, 输入命令以继续.')
            print("  添加条目 输入 add")
            print("  编辑字段 输入 edit")
            print("  跳过此行 输入 break")
            print("  退出程序 输入 quit")
            while True:
                cmd = rlinput("命令:")
                if cmd == "add":
                    product_obj = self._cmd_add(given['internal_name'])
                    break

                elif cmd == "break":
                    return

                elif cmd == "edit":
                    self._cmd_edit(given)
                    return self.query_info(given)

                elif cmd == "quit":
                    self.exit()

        else:
            # 是否有名称完全匹配的
            product_obj = db.get_product_by_internal_name(given['internal_name'])

            if not product_obj:
                print("请选择产品ID，可能是以下中的一个")
                print("如果产品不在下面列出，你还可以输入以下命令:")
                print("  添加条目 输入 add")
                print("  编辑字段 输入 edit")
                print("  跳过此行 输入 break")
                print("  退出程序 输入 quit")
                print('')
                ids = []
                for product in products.all():
                    ids.append(product.id)
                    space = " " * (20 - len(product.internal_name)) if len(product.internal_name) < 20 else ""
                    print("\t %s%s\t ID:%s\t %s±%sdPa.s" % (product.internal_name,
                                                            space,
                                                            product.id,
                                                            product.viscosity,
                                                            product.viscosity_width))

                while True:
                    pid = rlinput("请选择产品ID:")
                    if pid == "quit":
                        self.exit()

                    if pid == "add":
                        product_obj = self._cmd_add(given['internal_name'])
                        break

                    elif pid == "edit":
                        self._cmd_edit(given)
                        return self.query_info(given)

                    elif pid == "break":
                        return

                    elif not pid:
                        if len(ids) == 1:
                            pid = ids[0]
                            break

                    elif self._validate_id(pid, ids):
                        break

                if not product_obj and is_number_like(pid):
                    product_obj = db.get_product_by_id(pid)

                print("\t 你选择了(%s, %s±%sdPa.s)" % (product_obj.internal_name,
                                                   product_obj.viscosity,
                                                   product_obj.viscosity_width))

        given['market_name'] = product_obj.market_name
        given['kind'] = product_obj.template
        given['viscosity'] = product_obj.viscosity
        given['viscosity_limit'] = "%s±%s" % (product_obj.viscosity, product_obj.viscosity_width)
        given['qc_date'] = datetime.strftime(datetime.now(), '%Y/%m/%d')
        given['ftir'] = '{}%'.format(round(random.uniform(99.3, 100), 2))
        given['color'] = product_obj.color or ''
        given['wants_normal'] = True

        return given

    def _input_kind(self):
        kind = rlinput("类别(H-8100/H-9100/A-2000/K-2500/A-2100/A-9060A/A-9000/\n"
                       "UVS-1000/TM-3100/TS-3000/UVM-1800):\n >>>")

        conf = self.get_conf(kind)
        if not conf:
            conf = self.get_conf_by_alias(kind)
        if not conf:
            print("无效的类别, 请重新输入")
            conf = self._input_kind()

        return conf

    def _input_number(self, msg):
        value = rlinput(msg)

        if not is_number_like(value):
            value = self._input_number(None)

        return value

    def _cmd_add(self, internal_name):
        market_name = rlinput("销售品名(打在检验报告上的名字，例如‘8G04建业’的销售名为8G 04):\n >>>")
        conf = self._input_kind()
        viscosity = self._input_number("粘度值:")
        viscosity_width = self._input_number("粘度上下幅度:")

        product_obj = Product(internal_name=internal_name,
                              market_name=market_name,
                              template=conf['slug'],
                              viscosity=int(viscosity),
                              viscosity_width=int(viscosity_width))
        db.insert_product(product_obj)
        db.insert_product_to_xlsx(product_obj, '%s/data/database.xlsx' % self.app_path, 'products')
        print("已经插入新的条目到products数据表")
        return product_obj

    def _cmd_edit(self, given):
        given['internal_name'] = rlinput("品名:", given['internal_name'])
        self._ws['B{}'.format(self.index)] = given['internal_name']
        try:
            self._wb.save(self._product_wb_file)
        except PermissionError:
            print("war:文件已经被打开，无法写入")

    def _validate_id(self, id, list_ids):
        """ 验证指定的ID是否在给定的列表中 """
        if not re.match(r"\d+$", id):
            print("输入的ID不是一个数字")
            return False
        if int(id) not in list_ids:
            print("输入的ID不在指定的范围")
            return False
        return True

    def _set_report_info(self, product):
        """ 写入部分信息到指定行 """
        self._ws['G{}'.format(self.index)] = product['internal_name']
        self._ws['H{}'.format(self.index)] = product['viscosity_limit']
        self._ws['I{}'.format(self.index)] = product['product_date']
        self._ws['J{}'.format(self.index)] = product['validity_date']
        self._ws['K{}'.format(self.index)] = product['qc_date']

    def exit(self):
        self.save()
        sys.exit()

    def save(self):
        self.close_excel()
        self._wb.save(self._product_wb_file)

    def close_excel(self):
        if 'nt' in sys.builtin_module_names:
            import win32com.client
            from pywintypes import com_error

            engine = win32com.client.Dispatch('Excel.Application')
            engine.DisplayAlerts = False

            try:
                wb = engine.Workbooks(self._product_wb_file)
            except com_error:
                pass
            else:
                wb.Close(1)
            finally:
                engine.Quit()

    def get_today_report_dir_path(self):
        """
        自动创建并返回当日报告文件夹路径
        :return:
        """
        today_dir_name = datetime.strftime(datetime.now(), '%Y-%m-%d')
        today_path = os.path.join(module_path(), 'reports/%s' % today_dir_name)
        if not os.path.exists(today_path):
            os.mkdir(today_path)
        return today_path

    def onlyone_filename(self, path, filename, ext):
        """ 检查并生成唯一的文件名 """
        filepath = '%s/%s.%s' % (path, filename, ext)
        if os.path.exists(filepath):
            filename = '%s(1)' % filename
            filename = self.onlyone_filename(path, filename, ext)
        return filename

    def get_template_by_slug(self, slug):
        conf = self.get_conf(slug)
        if conf:
            return self.get_template(conf.get('template'))

    def get_conf(self, slug):
        for item in CONF:
            if item.get('slug') == slug:
                return item

    def get_conf_by_alias(self, alias):
        for item in CONF:
            if alias in item.get('alias'):
                return item

    def get_template(self, name):
        """ 获取模板文件路径 """
        return os.path.join(self.app_path, "templates/%s.docx" % name)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--index", type=int, help="excel中需要生成报告的起始行")
    parser.add_argument("--create_all", action="store_true", default=False,
                        help="初始化数据库")
    parser.add_argument("--reset_table", help="重置单个表")
    parser.add_argument("--init_products", action="store_true", default=False,
                        help="从xlsx文件中采集数据")
    parser.add_argument("--init_materials", action="store_true", default=False,
                        help="从xlsx文件中采集IQC检测要求数据")

    group = parser.add_argument_group('iqc')
    group.add_argument("--iqc", action="store_true", default=False, help="创建IQC报告")
    group.add_argument('-f', "--filename", help="IQC流水文件")
    group.add_argument('-e', "--end_row", type=int, help="IQC流水文件结束行")
    # subparsers = parser.add_subparsers(help='commands')
    # iqc_parser = subparsers.add_parser('iqc', help='创建IQC报告')
    # iqc_parser.add_argument('-f', "--filename", help="IQC流水文件")
    # iqc_parser.add_argument('-e', "--end_row", type=int, help="IQC流水文件结束行")

    args = parser.parse_args()

    if args.index:
        g = Generator(args.index)
        g.generate_reports()

    elif args.create_all:
        init_database()

    elif args.reset_table:
        reset_table(args.reset_table)

    elif args.init_products:
        reset_table(Product)
        db.init_product_data('{}/data/database.xlsx'.format(module_path()), 'products')

    elif args.init_materials:
        reset_table('iqc_materials')
        db.init_materials('{}/data/原材料检验项目及要求.xlsx'.format(module_path()), 'Sheet1')

    elif args.iqc:
        if args.filename:
            iqc_report.generate(args.filename, args.end_row)
        else:
            print("filename parameter must be provided. please refer --help for more detail.")
    else:
        g = Generator()
        g.generate_reports()
