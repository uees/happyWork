# -*- coding: utf-8 -*-
import random
import re
from datetime import datetime

from openpyxl import load_workbook

from common import module_path
from database import IQCMaterial


def generate(filename, end_row=None):
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('供应商来料质量统计表')
    if not end_row:
        end_row = ws.max_row
    for row in ws.iter_rows('B7:G{}'.format(end_row)):
        template_wb = load_workbook('%s/templates/iqc.xlsx' % module_path())
        template_ws = template_wb.get_sheet_by_name('Sheet1')
        material_name, incoming_date, supplier, qc_result, substandard_items, amount = [
            cell.value for cell in row]

        if isinstance(material_name, str):
            _material_name = re.sub(r'[\u4e00-\u9fa5]+', '', material_name)  # 去掉中文
            if not _material_name:
                continue
        material = IQCMaterial.query.filter(IQCMaterial.name.ilike('%' + _material_name + '%')).first()
        if not material or material.qc_items == '免检':
            continue

        if isinstance(incoming_date, datetime):
            incoming_date = datetime.strftime(incoming_date, '%Y-%m-%d')  # 转为字符串

        if _material_name.upper() in ['0.25L', '0.3L', '1L', '5L', '6L', '20L',
                                      '0.25KG', '0.3KG', '1KG', '5KG', '6KG',
                                      '20KG']:
            unit = '套'
        else:
            unit = 'kg'

        template_ws.cell('B5').value = incoming_date
        template_ws.cell('B6').value = material_name
        template_ws.cell('D6').value = supplier
        template_ws.cell('D7').value = '%s%s' % (amount, unit)
        template_ws.cell('D8').value = material.qc_method

        qc_items = material.qc_items.split('、')
        row = 11
        for item in qc_items:
            template_ws.cell('A{}'.format(row)).value = item
            if item == '细度':
                if _material_name.find('A0084') >= 0:
                    template_ws.cell('C{}'.format(row)).value = '<25μm'
                elif _material_name.find('A0085') >= 0 or \
                        _material_name.find('A0088') >= 0:
                    template_ws.cell('C{}'.format(row)).value = '<17.5μm'
                else:
                    template_ws.cell('C{}'.format(row)).value = '<20μm'
            elif item == '软化点':
                if _material_name == 'A0016' or _material_name == 'A0016A'\
                        or _material_name == 'A0016B':
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s℃' % round(random.uniform(27, 30), 1)
                elif _material_name == 'A0016F':
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s℃' % round(random.uniform(32, 35), 1)
                else:
                    template_ws.cell('C{}'.format(row)).value = '√'
            elif item == '环氧值':
                if _material_name == 'A0016' or _material_name == 'A0016A'\
                        or _material_name == 'A0016B':
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s mol/100g' % round(random.uniform(0.515, 0.535), 3)
                elif _material_name == 'A0016F':
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s mol/100g' % round(random.uniform(0.56, 0.59), 3)
                else:
                    template_ws.cell('C{}'.format(row)).value = '√'
            elif item == '馏程':
                if _material_name.find('A0055') >= 0 or \
                        _material_name.find('A0063') >= 0:
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s~%s℃' % (str(180 + random.randint(1, 9)),
                                    str(220 - random.randint(1, 9)))
                elif _material_name.find('A0058') >= 0:
                    template_ws.cell('C{}'.format(row)).value = \
                        '%s~%s℃' % (str(135 + random.randint(1, 5)),
                                    str(150 - random.randint(1, 5)))
                else:
                    template_ws.cell('C{}'.format(row)).value = '√'
            else:
                template_ws.cell('C{}'.format(row)).value = '√'
            row += 1
        template_ws.merge_cells('B11:B{}'.format(10 + len(qc_items)))
        template_ws.cell('B11').value = material.spec

        new_filename = '%s-%s-%s-%s.xlsx' % (incoming_date,
                                             random.randint(1, 99),
                                             material_name,
                                             supplier)
        template_wb.save('%s/reports/IQC/%s' % (module_path(), new_filename))
