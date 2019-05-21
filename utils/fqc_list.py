# -*- coding: utf-8 -*-
import random

from openpyxl import load_workbook
from openpyxl.styles import Font

from settings import FQC_ITEMS, ALL_FQC_ITEMS


class FqcListGenerator(object):

    def __init__(self, fqc_filename):
        self.fqc_filename = fqc_filename
        self.fqc_wb = load_workbook(fqc_filename)
        self.fqc_ws = self.fqc_wb.get_sheet_by_name('FQC')
        self.start_row = self.fqc_ws.max_row
        self.added = 0

    def save(self):
        self.fqc_wb.save(self.fqc_filename)

    def fqc_record(self, product):
        _font = Font(name='Calibri', size=10)
        row = self.start_row + self.added
        record = self._make_record(product)

        if record:
            # 设置前4列
            for (i, item) in zip(range(1, 5), [
                product['product_date'],
                product['qc_date'],
                product['internal_name'],
                product['batch']
            ]):
                self.fqc_ws.cell(row=row, column=i).value = item
                self.fqc_ws.cell(row=row, column=i).font = _font

            # 设置记录值
            for item in record:
                self.fqc_ws.cell(row=row, column=item['col']).value = item['value']
                self.fqc_ws.cell(row=row, column=item['col']).font = _font

        self.added += 1

    def _make_record(self, product):
        c = product["kind"]
        items = FQC_ITEMS.get(c)
        if not items:
            return

        record = []
        for item in items:
            if item == '外观颜色' or item == '板面效果' or item == '固化性'\
                    or item == '显影性' or item == '去膜性'\
                    or item == '耐焊性' or item == '耐化学性':
                record.append({'col': self._get_col(item), 'value': '√'})
            elif item == '细度':
                record.append({'col': self._get_col(item),
                               'value': '%sμm' % random.choice([15, 17.5, 20])})
            elif item == '反白条':
                if c == 'uvw5d10' or c == 'uvw5d65' or c == 'a2' or c == 'k2'\
                        or c == 'thw5d35':
                    record.append(
                        {'col': self._get_col(item), 'value': '≤10μm'})
                elif c == 'tm3100' or c == 'ts3000' or c == 'uvs1000' or c == 'uvm1800':
                    record.append(
                        {'col': self._get_col(item), 'value': '≤20μm'})
                else:
                    record.append(
                        {'col': self._get_col(item), 'value': '≤5μm'})
            elif item == '粘度':
                if product['internal_name'].find('A-9060C明阳') >= 0:
                    unit = 's/32℃'
                else:
                    unit = 'dpa.s/25℃'
                record.append({'col': self._get_col(item),
                               'value': '%s%s' % (self._get_viscosity(product['viscosity']), unit)})
            elif item == '硬度':
                if c == 'uvw5d10':
                    record.append({'col': self._get_col(item), 'value': 'H'})
                elif c == 'a9':
                    record.append({'col': self._get_col(item), 'value': '2H'})
                elif c == 'uvm1800':
                    record.append({'col': self._get_col(item), 'value': '3H'})
                elif c == 'tm3100' or c == 'uvs1000':
                    record.append({'col': self._get_col(item), 'value': '4H'})
                elif c == 'ts3000':
                    record.append({'col': self._get_col(item), 'value': '5H'})
                elif c == 'h9100' or c == 'h8100':
                    record.append({'col': self._get_col(item), 'value': '6H'})
            elif item == '附着力':
                record.append({'col': self._get_col(item), 'value': '100%'})
            elif item == '感光性':
                if c == 'h9100' or c == 'h8100':
                    record.append({'col': self._get_col(item),
                                   'value': '%s级' % random.choice([9.5, 10, 10.5, 11])})
                elif c == 'a9':
                    record.append({'col': self._get_col(item),
                                   'value': '%s级' % random.choice([6, 6.5, 7])})
            elif item == '解像性':
                record.append({'col': self._get_col(item), 'value': '≤50μm'})
            elif item == '红外图谱':
                record.append({'col': self._get_col(item),
                               'value': '{}%'.format(round(random.uniform(99.3, 100), 2))})
        return record

    def _get_viscosity(self, viscosity):
        viscosity = int(viscosity)
        if viscosity <= 30:
            viscosity += random.uniform(-1, 1)
        elif viscosity <= 100:
            viscosity += random.uniform(-5, 5)
        else:
            viscosity += random.uniform(-10, 10)
        return round(viscosity, 1)

    def _get_col(self, item):
        col = ALL_FQC_ITEMS.index(item) + 5
        return col
