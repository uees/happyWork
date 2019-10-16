import math
import re
import sys

from openpyxl import load_workbook

from common import rlinput
from package_material.models.consts import COL_INDEXES, KIND_PACKAGES, PACKAGE_CATEGORIES
from package_material.models.product import Product


class WorksheetParser(object):
    """
    解析一个工作表中的包材信息
    """

    def __init__(self, filename, sheet, start_row=2):
        self.filename = filename
        self.sheet = sheet
        self.start_row = start_row
        self.wb = load_workbook(filename)
        self.ws = self.wb[sheet]

    def run(self):
        current_row = self.start_row
        for row in self.ws[f'A{self.start_row}:J{self.ws.max_row}']:
            _type, date, NO, custmor, code, name, spec, batch, weight, made_at = row

            # 无品名 和 重量负数 的不统计
            # 不是 产品进仓 和 改标进仓 的不统计
            if not name.value or weight.value < 0 \
                    or (_type.value != "产品进仓" and _type.value != "改标进仓"):
                current_row += 1
                continue

            origin_name = name.value
            print(f"正在计算第{current_row}行 {origin_name} {batch.value} {weight.value}kg 的包材用量")

            product = self.query_product(origin_name)
            if not product:
                current_row += 1
                continue

            # 获取单箱重量
            per_weight = self.get_per_weight(spec.value)
            if not per_weight:
                current_row += 1
                continue

            try:
                category = self.get_package_category(product, per_weight, origin_name)
            except KeyError as e:
                print(e)
                self.wb.save(self.filename)
                sys.exit()

            # fix 单罐重量
            if per_weight < 2:
                # 1kg 包装 10 罐 / 箱
                per_weight *= 10
            elif per_weight < 5:
                # 5kg 包装 4 罐 / 箱
                per_weight *= 4

            # 箱数，20L 罐包装就是是罐数
            amount = weight.value / per_weight  # float
            box_type = category['box_type']
            box_amount = category['box_amount']
            part_a_jar_type = category['part_a_jar_type']
            part_a_jar_amount = category['part_a_jar_amount']
            part_b_jar_type = category['part_b_jar_type']
            part_b_jar_amount = category['part_b_jar_amount']
            label_amount = category['label_amount']

            # 标签用量
            # math.ceil 向上取整 2.3 -> 3
            self.ws[f'Y{current_row}'] = math.ceil(label_amount * amount)

            # 纸箱用量
            if box_type:
                self.ws[f'{COL_INDEXES[box_type]}{current_row}'] = math.ceil(box_amount * amount)

            # 改标进仓不消耗罐子
            if _type.value != "改标进仓":
                if part_a_jar_type:
                    self.ws[f'{COL_INDEXES[part_a_jar_type]}{current_row}'] = math.ceil(part_a_jar_amount * amount)
                if part_b_jar_type:
                    self.ws[f'{COL_INDEXES[part_b_jar_type]}{current_row}'] = math.ceil(part_b_jar_amount * amount)

            current_row += 1

        self.wb.save(self.filename)
        print("计算完毕")

    def query_product(self, product_name):
        """
        :param product_name:
        :return: if None 表示 break
        """
        # 光刻胶
        if product_name.endswith("CP") \
                or product_name.startswith("RDR-") \
                or product_name.startswith("RD-") \
                or product_name.startswith("RDJ-") \
                or product_name.startswith("SIJ-"):
            return

        # 开油水
        if product_name.startswith("S-") \
                or product_name.find('助剂') >= 0 \
                or product_name.endswith("固化剂"):  # tx-1109 固化剂
            return

        # 优化 product_name
        if product_name.startswith("9GHD"):
            product_name = "9G"
        elif product_name.startswith("2GHD"):
            product_name = "2G"
        elif product_name.startswith("3GHD"):
            product_name = "3G"
        elif product_name.startswith("8G04HD"):
            product_name = "8G04"
        elif product_name.startswith("UVS-1000"):
            product_name = "UVS-1000"

        product_name = product_name \
            .replace("内袋", "") \
            .replace("固内", "") \
            .replace("胜宏", "") \
            .replace("金像", "") \
            .replace("川亿", "") \
            .replace("外贸", "") \
            .strip()

        while True:
            products = Product.query.search(product_name).limit(20).all()

            if products:
                break

            print(f"未查到{product_name}的产品记录，请修改一下")
            product_name = rlinput("品名:", product_name)
            # product_name = rlinput("品名:")
            if product_name == 'break' or product_name == 'b':
                return

            elif product_name == "quit":
                self.wb.save(self.filename)
                sys.exit()

        if products:
            return self.select_product(products, product_name)

    def select_product(self, products, product_name):
        """
        :param products: 产品列表
        :param product_name: 产品名称
        :return: if None 表示 break
        """
        for product in products:
            if product.internal_name == product_name:
                return product

        print(f"请选择产品{product_name}的ID，可能是以下中的一个")
        for product in products:
            space = " " * (20 - len(product.internal_name)) if len(product.internal_name) < 20 else ""
            print("\t %s%s\t ID:%s" % (product.internal_name, space, product.id))

        while True:
            pid = rlinput("请选择产品ID:")
            if pid == "quit":
                self.wb.save(self.filename)
                sys.exit()

            elif pid == "break" or pid == 'b':
                return

            # 直接回车就是选择第一个
            elif not pid:
                return products[0]

            for product in products:
                if str(product.id) == pid:
                    return product

            print(f"无效的id: {pid}")

    @staticmethod
    def get_package_category(product, per_weight, origin_name):
        slug = product.category.slug

        # 未分类的都是单组份油墨，类似 UVS-1000 的包装
        if slug == 'undefined':
            slug = "UVS-1000"

        kind = KIND_PACKAGES[slug]

        # fix 单罐重量
        if per_weight < 2:
            return PACKAGE_CATEGORIES[kind['10kg']]
        if per_weight < 5:
            return PACKAGE_CATEGORIES[kind['20kg']]

        # 低压喷涂油和静电喷涂油
        if origin_name.find('SP') >= 0:
            kind = KIND_PACKAGES['H-9100 SP']
            if origin_name.find('内袋') >= 0:
                return PACKAGE_CATEGORIES[kind['20kg内袋']]
            return PACKAGE_CATEGORIES[kind['20kg']]

        # 感光阻焊油
        if slug == 'H-8100' or slug == 'H-9100':
            if per_weight == 5:
                return PACKAGE_CATEGORIES[kind['5kg']]
            if per_weight == 25:
                return PACKAGE_CATEGORIES[kind['25kg内袋']]

        if per_weight <= 10:
            key = '10kg'
        else:
            key = '20kg'

        if origin_name.find('内袋') >= 0:
            key += '内袋'
        elif origin_name.find('固内') >= 0:
            key += '固内'
        return PACKAGE_CATEGORIES[kind[key]]

    @staticmethod
    def get_per_weight(spec):
        match = re.match(r'^\d+\.?\d+', spec)
        if match:
            per_weight = float(match.group())
            return per_weight
        return 0
