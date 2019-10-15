import argparse
import re
import os
import sys

import requests
from openpyxl import load_workbook

from common import rlinput
from package_material.models.category import KIND_PACKAGES, PACKAGE_CATEGORIES, COL_INDEXES


def login(host, email, password):
    response = requests.post(f'{host}/api/auth/login', data={
        'email': email, 'password': password
    }, headers={
        'X-Requested-With': 'XMLHttpRequest'
    })

    body = response.json()

    return body.get('token')


def search_product(host, token, name):
    response = requests.get(f'{host}/api/products', params={
        "with": "category,testWay",
        "page": 1,
        "per_page": 40,
        "sort_by": "id",
        "order": "desc",
        "q": name
    }, headers={
        'X-Requested-With': 'XMLHttpRequest',
        'Authorization': f'Bearer {token}'
    })

    products = response.json().get('data')

    return products


def select_product(products, product_name):
    for product in products:
        if product['internal_name'] == product_name:
            return product

    print("请选择产品ID，可能是以下中的一个")
    for product in products:
        space = " " * (20 - len(product['internal_name'])) if len(product['internal_name']) < 20 else ""
        print("\t %s%s\t ID:%s" % (product['internal_name'], space, product['id']))

    while True:
        pid = rlinput("请选择产品ID:")
        if pid == "quit":
            sys.exit()

        elif pid == "break":
            return

        for product in products:
            if str(product['id']) == pid:
                print("hehe")
                return product

        print(f"无效的id: {pid}")


def get_package_category(product, per_weight, origin_name):
    slug = product.get('category').get('slug')
    kind = KIND_PACKAGES[slug]

    # 固化剂是静电喷涂的
    if slug == 'H-8100B/H-9100B':
        return PACKAGE_CATEGORIES[kind['SP']]

    # 低压喷涂油和静电喷涂油
    if origin_name.find('SP') >= 0:
        kind = KIND_PACKAGES['H-9100 SP']
        if origin_name.find('内袋') >= 0:
            return PACKAGE_CATEGORIES[kind['20kg内袋']]
        return PACKAGE_CATEGORIES[kind['20kg']]

    if per_weight == 5 and (slug == 'H-8100' or slug == 'H-9100'):
        return PACKAGE_CATEGORIES[kind['5kg']]
    elif per_weight < 10:
        key = '10kg'
    else:
        key = '20kg'

    if origin_name.find('内袋') >= 0:
        key += '内袋'
    elif origin_name.find('固内') >= 0:
        key += '固内'
    return PACKAGE_CATEGORIES[kind[key]]


def get_per_weight(spec):
    match = re.match(r'^\d+\.?\d+', spec)
    if match:
        per_weight = float(match.group())
        return per_weight
    return 0


def load_file(filename, sheet="产品进仓", start_row=2):
    host = os.getenv('QC_HOST')
    email = os.getenv('ADMIN_USER')
    passwd = os.getenv('ADMIN_PASS')
    token = login(host, email, passwd)

    # read_only 可防止内存爆出，data_only 可以读取公式的值，而不是读到公式
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet]
    current_row = start_row
    for row in ws[f'A{start_row}:J{ws.max_row}']:
        _type, date, NO, custmor, code, name, spec, batch, weight, made_at = row

        if not name.value:
            continue

        origin_name = name.value
        product_name = origin_name
        products = search_product(host, token, product_name)

        while not products:
            product_name = rlinput("品名:", product_name)
            if product_name == 'break':
                break

            if product_name == "quit":
                sys.exit()

            products = search_product(host, token, product_name)

        product = select_product(products, product_name)

        per_weight = get_per_weight(spec.value)
        if not per_weight:
            break

        category = get_package_category(product, per_weight, origin_name)

        amount = int(weight.value / per_weight)
        box_type = category['box_type']
        box_amount = category['box_amount']
        part_a_jar_type = category['part_a_jar_type']
        part_a_jar_amount = category['part_a_jar_amount']
        part_b_jar_type = category['part_b_jar_type']
        part_b_jar_amount = category['part_b_jar_amount']
        c_weight = category['weight']
        label_amount = category['label_amount']

        if box_type:
            ws['{}{}'.format(COL_INDEXES[box_type], current_row)] = box_amount * amount
        if part_a_jar_type:
            ws['{}{}'.format(COL_INDEXES[part_a_jar_type], current_row)] = part_a_jar_amount * amount
        if part_b_jar_type:
            ws['{}{}'.format(COL_INDEXES[part_b_jar_type], current_row)] = part_b_jar_amount * amount
        ws['Y{}'.format(current_row)] = label_amount * amount

        current_row += 1

    wb.save(filename)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='分析包材用量')
    parser.add_argument("excel_file", help="文件名")
    parser.add_argument("-i", "--index", type=int, default=2, help="excel中需要生成报告的起始行")
    parser.add_argument("-s", "--sheet", default="产品进仓", help="excel中的工作表")

    args = parser.parse_args()
    load_file(args.excel_file, args.sheet, args.index)
